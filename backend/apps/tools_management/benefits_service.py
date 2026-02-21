from collections import defaultdict
from datetime import date

from openpyxl import load_workbook


DEZHOU_REQUIRED_COLS = {'俱乐部', '玩家昵称', '盲注', '贵宾分保险', '本局手数'}
NIUZAI_REQUIRED_COLS = {'俱乐部', '玩家昵称', '战绩', '下注量'}

CACHE_KEY_BENEFITS = 'tools_benefits_data'
CACHE_TTL_BENEFITS = 60 * 60 * 24 * 90


def get_benefits_cache_key(query_date=''):
    day = (query_date or date.today().isoformat()).strip()
    return f'{CACHE_KEY_BENEFITS}:{day}'


def analyze_benefits_files(files):
    dezhou_rows = []
    niuzai_rows = []

    for file_obj in files:
        filename = (getattr(file_obj, 'name', '') or '').lower()
        if not filename.endswith('.xlsx'):
            raise ValueError(f'仅支持 xlsx 文件: {getattr(file_obj, "name", "")}')

        workbook = load_workbook(file_obj, data_only=True)
        sheet = workbook.active
        headers, header_idx = _find_header_row(sheet)
        if not headers:
            continue

        header_set = set(headers)
        rows = _read_rows(sheet, headers, header_idx)
        if DEZHOU_REQUIRED_COLS.issubset(header_set):
            dezhou_rows.extend(rows)
        if NIUZAI_REQUIRED_COLS.issubset(header_set):
            niuzai_rows.extend(rows)

    return build_benefits_stats(dezhou_rows, niuzai_rows)


def build_benefits_stats(dezhou_rows, niuzai_rows):
    vip_sum = defaultdict(float)
    blind_counts = defaultdict(int)
    cowboy_sum = defaultdict(float)
    cowboy_bet_sum = defaultdict(float)
    players_by_club = defaultdict(set)

    for row in dezhou_rows:
        club = _to_text(row.get('俱乐部'))
        player = _to_text(row.get('玩家昵称'))
        if not club or not player:
            continue
        players_by_club[club].add(player)
        vip_sum[(club, player)] += _to_float(row.get('贵宾分保险'))

        hands = _to_float(row.get('本局手数'))
        if hands >= 30:
            blind = _to_text(row.get('盲注')) or '未知盲注'
            blind_counts[(club, player, blind)] += 1

    for row in niuzai_rows:
        club = _to_text(row.get('俱乐部'))
        player = _to_text(row.get('玩家昵称'))
        if not club or not player:
            continue
        players_by_club[club].add(player)
        cowboy_sum[(club, player)] += _to_float(row.get('战绩'))
        cowboy_bet_sum[(club, player)] += _to_float(row.get('下注量'))

    result = {}
    for club in sorted(players_by_club.keys()):
        club_map = {}
        for player in sorted(players_by_club[club]):
            key = (club, player)
            player_data = {
                '玩家贵宾分保险总数': round(vip_sum.get(key, 0.0), 2),
                '牛仔战绩': round(cowboy_sum.get(key, 0.0), 2),
                '牛仔下注量汇总': round(cowboy_bet_sum.get(key, 0.0), 2)
            }

            blind_total = 0
            blind_map = {}
            for (bc_club, bc_player, blind), count in blind_counts.items():
                if bc_club == club and bc_player == player:
                    blind_map[blind] = count
                    blind_total += count

            player_data['玩家打满30次数'] = int(blind_total)
            for blind in sorted(blind_map.keys()):
                player_data[blind] = int(blind_map[blind])

            club_map[player] = player_data
        result[club] = club_map
    return result


def _find_header_row(sheet):
    for idx in (1, 2, 3):
        values = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[idx]]
        values = [v for v in values if v]
        if not values:
            continue
        value_set = set(values)
        if DEZHOU_REQUIRED_COLS.issubset(value_set) or NIUZAI_REQUIRED_COLS.issubset(value_set):
            return values, idx
    return [], 0


def _read_rows(sheet, headers, header_idx):
    rows = []
    header_len = len(headers)
    for values in sheet.iter_rows(min_row=header_idx + 1, values_only=True):
        row = {}
        non_empty = False
        for i in range(header_len):
            key = headers[i]
            value = values[i] if i < len(values) else None
            row[key] = value
            if value not in (None, ''):
                non_empty = True
        if non_empty:
            rows.append(row)
    return rows


def _to_float(value):
    if value is None or value == '':
        return 0.0
    try:
        return float(str(value).replace(',', '').strip())
    except Exception:
        return 0.0


def _to_text(value):
    if value is None:
        return ''
    return str(value).strip()
