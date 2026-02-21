import base64
import io
import re
from django.db import transaction
from openpyxl import load_workbook, Workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from apps.core.decorators import permission_required
from apps.core.permissions import HasPermission
from apps.core.response import success_response, error_response
from apps.script.models import ScriptQueue
from apps.script.redis_client import get_redis
from .models import ArchivePhoneData
from .serializers import ArchivePhoneDataSerializer


PHONE_PATTERN = re.compile(r'^1[3-9]\d{9}$')
EXPECTED_IMPORT_HEADERS = ['phone', 'source', 'remark', 'platforms', 'is_filtered']
PLATFORM_META = {
    1 << 0: {'code': 'wechat', 'label': '微信'},
    1 << 1: {'code': 'alipay', 'label': '支付宝'},
    1 << 2: {'code': 'douyin', 'label': '抖音'},
    1 << 3: {'code': 'qq', 'label': 'QQ'},
    1 << 4: {'code': 'santiao', 'label': '三条'}
}
PLATFORM_NAME_TO_BIT = {meta['code']: bit for bit, meta in PLATFORM_META.items()}


def _parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).lower() in ['1', 'true', 'yes', 'on']


def _platforms_to_flags(platforms):
    flags = 0
    for platform in platforms:
        bit = PLATFORM_NAME_TO_BIT.get((platform or '').lower())
        if bit:
            flags |= bit
    return flags


def _flags_to_platforms(flags):
    return ','.join([meta['code'] for bit, meta in PLATFORM_META.items() if flags & bit])


def _split_platforms(raw):
    if not raw:
        return []
    if isinstance(raw, list):
        return [str(item).strip().lower() for item in raw if str(item).strip()]
    text = str(raw).replace('|', ',').replace(';', ',')
    return [item.strip().lower() for item in text.split(',') if item.strip()]


def _append_archive_records(rows):
    if not rows:
        return {'total': 0, 'created': 0, 'skipped': 0, 'invalid': 0}

    dedup = {}
    invalid = 0
    duplicate = 0
    failed_samples = []
    for row in rows:
        phone = str(row.get('phone') or '').strip()
        if not PHONE_PATTERN.match(phone):
            invalid += 1
            if len(failed_samples) < 100:
                failed_samples.append({
                    'line': row.get('_line', 0),
                    'phone': phone,
                    'reason': '手机号格式不正确'
                })
            continue
        if phone in dedup:
            duplicate += 1
            if len(failed_samples) < 100:
                failed_samples.append({
                    'line': row.get('_line', 0),
                    'phone': phone,
                    'reason': '文件内重复手机号'
                })
            continue
        dedup[phone] = row

    if not dedup:
        return {'total': 0, 'created': 0, 'skipped': 0, 'invalid': invalid, 'duplicate': duplicate, 'failed_samples': failed_samples}

    phones = [int(phone) for phone in dedup.keys()]
    existing_map = {
        item.phone: item
        for item in ArchivePhoneData.objects.filter(phone__in=phones)
    }

    creates = []
    skipped = 0

    for phone_str, row in dedup.items():
        phone = int(phone_str)
        platforms = _split_platforms(row.get('platforms'))
        flags = _platforms_to_flags(platforms)
        data = {
            'phone': phone,
            'is_filtered': _parse_bool(row.get('is_filtered'), default=False),
            'status': 1,
            'source': (row.get('source') or 'manual').strip() or 'manual',
            'register_flags': flags,
            'platforms': _flags_to_platforms(flags),
            'remark': (row.get('remark') or '').strip(),
            'extra': row.get('extra') or {}
        }

        obj = existing_map.get(phone)
        if obj:
            skipped += 1
            if len(failed_samples) < 100:
                failed_samples.append({
                    'line': row.get('_line', 0),
                    'phone': str(phone),
                    'reason': '数据库已存在，已跳过'
                })
        else:
            creates.append(ArchivePhoneData(**data))

    if creates:
        ArchivePhoneData.objects.bulk_create(creates, batch_size=1000)

    return {
        'total': len(dedup),
        'created': len(creates),
        'skipped': skipped,
        'invalid': invalid,
        'duplicate': duplicate,
        'failed_samples': failed_samples
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:list')
def archive_data_list_view(request):
    page = int(request.GET.get('page', 1) or 1)
    limit = int(request.GET.get('limit', 20) or 20)
    page = max(page, 1)
    limit = max(1, min(limit, 100))

    queryset = ArchivePhoneData.objects.filter(is_deleted=False, status=1).order_by('-updated_at', '-id')

    phone = (request.GET.get('phone') or '').strip()
    remark = (request.GET.get('remark') or '').strip()
    source = (request.GET.get('source') or '').strip()
    is_filtered = request.GET.get('is_filtered')
    platforms = request.GET.getlist('platforms') or request.GET.getlist('platforms[]')

    if phone:
        queryset = queryset.filter(phone=phone) if phone.isdigit() else queryset.none()
    if remark:
        queryset = queryset.filter(remark__icontains=remark)
    if source:
        queryset = queryset.filter(source=source)
    if is_filtered in ['0', '1']:
        queryset = queryset.filter(is_filtered=(is_filtered == '1'))
    if platforms:
        for platform in platforms:
            queryset = queryset.filter(platforms__contains=platform)

    total = queryset.count()
    start = (page - 1) * limit
    end = start + limit
    rows = queryset[start:end]

    return success_response({
        'items': ArchivePhoneDataSerializer(rows, many=True).data,
        'total': total,
        'page': page,
        'limit': limit
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:list')
def archive_data_platform_meta_view(request):
    return success_response(list(PLATFORM_META.values()))


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:add')
def archive_data_create_view(request):
    phone = str(request.data.get('phone') or '').strip()
    if not PHONE_PATTERN.match(phone):
        return error_response('手机号格式不正确', code=400, http_status=status.HTTP_400_BAD_REQUEST)

    platforms = request.data.get('platforms') or []
    if not isinstance(platforms, list):
        platforms = [platforms]

    flags = _platforms_to_flags(platforms)
    payload = {
        'phone': int(phone),
        'is_filtered': _parse_bool(request.data.get('is_filtered'), default=False),
        'status': 1,
        'source': (request.data.get('source') or 'manual').strip() or 'manual',
        'register_flags': flags,
        'platforms': _flags_to_platforms(flags),
        'remark': (request.data.get('remark') or '').strip(),
        'extra': request.data.get('extra') or {}
    }

    existing = ArchivePhoneData.objects.filter(phone=payload['phone']).first()
    if existing:
        existing.is_deleted = False
        for key, value in payload.items():
            setattr(existing, key, value)
        existing.save()
        return success_response(ArchivePhoneDataSerializer(existing).data, message='保存成功')

    instance = ArchivePhoneData.objects.create(**payload)
    return success_response(
        ArchivePhoneDataSerializer(instance).data,
        message='创建成功',
        code=201,
        http_status=status.HTTP_201_CREATED
    )


@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:edit')
def archive_data_update_view(request, data_id):
    try:
        instance = ArchivePhoneData.objects.get(id=data_id, is_deleted=False)
    except ArchivePhoneData.DoesNotExist:
        return error_response('数据不存在', code=404, http_status=status.HTTP_404_NOT_FOUND)

    phone = request.data.get('phone')
    if phone is not None:
        phone = str(phone).strip()
        if not PHONE_PATTERN.match(phone):
            return error_response('手机号格式不正确', code=400, http_status=status.HTTP_400_BAD_REQUEST)
        instance.phone = int(phone)

    if 'platforms' in request.data:
        platforms = request.data.get('platforms') or []
        if not isinstance(platforms, list):
            platforms = [platforms]
        flags = _platforms_to_flags(platforms)
        instance.register_flags = flags
        instance.platforms = _flags_to_platforms(flags)

    if 'is_filtered' in request.data:
        instance.is_filtered = _parse_bool(request.data.get('is_filtered'), default=False)
    if 'source' in request.data:
        instance.source = (request.data.get('source') or 'manual').strip() or 'manual'
    if 'remark' in request.data:
        instance.remark = (request.data.get('remark') or '').strip()
    if 'extra' in request.data:
        instance.extra = request.data.get('extra') or {}

    instance.save()
    return success_response(ArchivePhoneDataSerializer(instance).data, message='更新成功')


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:delete')
def archive_data_delete_view(request, data_id):
    try:
        instance = ArchivePhoneData.objects.get(id=data_id, is_deleted=False)
    except ArchivePhoneData.DoesNotExist:
        return error_response('数据不存在', code=404, http_status=status.HTTP_404_NOT_FOUND)

    instance.is_deleted = True
    instance.save(update_fields=['is_deleted'])
    return success_response(message='删除成功')


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:import')
def archive_data_import_view(request):
    file = request.FILES.get('file')
    if not file:
        return error_response('请上传文件', code=400, http_status=status.HTTP_400_BAD_REQUEST)

    filename = (file.name or '').lower()
    if not filename.endswith('.xlsx'):
        return error_response('仅支持导入 XLSX 模板文件', code=400, http_status=status.HTTP_400_BAD_REQUEST)

    rows = []
    try:
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return error_response('XLSX 文件解析失败', code=400, http_status=status.HTTP_400_BAD_REQUEST)

    if not all_rows:
        return error_response('导入文件为空', code=400, http_status=status.HTTP_400_BAD_REQUEST)

    header = [str(item).strip() if item is not None else '' for item in all_rows[0]]
    if header != EXPECTED_IMPORT_HEADERS:
        return error_response(
            f"模板列名不匹配，必须严格为: {','.join(EXPECTED_IMPORT_HEADERS)}",
            code=400,
            http_status=status.HTTP_400_BAD_REQUEST
        )

    for idx, row in enumerate(all_rows[1:], start=2):
        row_values = list(row) if row is not None else []
        if not row_values or not ''.join([str(item).strip() for item in row_values if item is not None]):
            continue
        if len(row_values) != len(EXPECTED_IMPORT_HEADERS):
            return error_response(
                f'第 {idx} 行列数不正确，必须为 {len(EXPECTED_IMPORT_HEADERS)} 列',
                code=400,
                http_status=status.HTTP_400_BAD_REQUEST
            )

        rows.append({
            '_line': idx,
            'phone': str(row_values[0]).strip() if row_values[0] is not None else '',
            'source': (str(row_values[1]).strip() if row_values[1] is not None else '') or 'import',
            'remark': str(row_values[2]).strip() if row_values[2] is not None else '',
            'platforms': str(row_values[3]).strip() if row_values[3] is not None else '',
            'is_filtered': str(row_values[4]).strip() if row_values[4] is not None else ''
        })

    with transaction.atomic():
        result = _append_archive_records(rows)

    return success_response(result, message='导入完成')


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:import')
def archive_data_import_template_view(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'archive_import'
    ws.append(EXPECTED_IMPORT_HEADERS)
    ws.append(['13800138000', 'manual', '示例备注', 'wechat|alipay', '0'])
    ws.append(['13900139000', 'import', '批量导入样例', 'santiao|qq', '1'])

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    encoded = base64.b64encode(buffer.getvalue()).decode('utf-8')

    return success_response({
        'filename': 'archive_import_template.xlsx',
        'content_base64': encoded,
        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'note': 'platforms 支持英文平台代码，使用 | 或 , 分隔'
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasPermission])
@permission_required('data:archive:sync')
def archive_data_sync_view(request):
    limit_per_queue = int(request.data.get('limit_per_queue', 5000) or 5000)
    limit_per_queue = max(1, min(limit_per_queue, 50000))

    queues = ScriptQueue.objects.filter(is_deleted=False)
    r = get_redis()
    merged = []
    for queue in queues:
        key = f'queue:{queue.name}'
        values = r.lrange(key, 0, limit_per_queue - 1)
        for value in values:
            phone = value.decode('utf-8') if isinstance(value, bytes) else str(value)
            merged.append({
                'phone': phone.strip(),
                'source': f'queue:{queue.name}',
                'remark': '',
                'platforms': ''
            })

    with transaction.atomic():
        result = _append_archive_records(merged)
    result['queues'] = queues.count()
    result['scanned'] = len(merged)
    return success_response(result, message='同步完成')
